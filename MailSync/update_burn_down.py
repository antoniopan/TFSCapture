import openpyxl
import openpyxl.utils.datetime
import sys
import datetime
import math


def update_burn_down_xlsx(src, dst):
    wb_module = openpyxl.load_workbook(src)
    ws_module = wb_module['Task Table']

    wb_bd = openpyxl.load_workbook(dst)
    ws_sr = wb_bd['SolvedRate']
    ws_d = wb_bd['解决率距离']
    ws_o = wb_bd['Open']

    row_sr = match_today(ws_sr)
    row_d = match_today(ws_d)
    row_o = match_today(ws_o)

    for i in range(1, ws_module.max_row + 1):
        module = ws_module.cell(i, 1).value
        sr = ws_module.cell(i, 6).value
        d = math.ceil(ws_module.cell(i,5).value*0.9-ws_module.cell(i,3).value-ws_module.cell(i,4).value)
        open = ws_module.cell(i, 2).value
        if '3D' == module:
            ws_sr.cell(row_sr, 2).value = sr
            ws_d.cell(row_d, 2).value = d
            ws_o.cell(row_o, 2).value = open
        elif 'Exam' == module:
            ws_sr.cell(row_sr, 3).value = sr
            ws_d.cell(row_d, 3).value = d
            ws_o.cell(row_o, 3).value = open
        elif 'Filming' == module:
            ws_sr.cell(row_sr, 4).value = sr
            ws_d.cell(row_d, 4).value = d
            ws_o.cell(row_o, 4).value = open
        elif 'MainFrame' == module:
            ws_sr.cell(row_sr, 5).value = sr
            ws_d.cell(row_d, 5).value = d
            ws_o.cell(row_o, 5).value = open
        elif 'PA' == module:
            ws_sr.cell(row_sr, 6).value = sr
            ws_d.cell(row_d, 6).value = d
            ws_o.cell(row_o, 6).value = open
        elif 'PR' == module:
            ws_sr.cell(row_sr, 7).value = sr
            ws_d.cell(row_d, 7).value = d
            ws_o.cell(row_o, 7).value = open
        elif 'Ref' == module:
            ws_sr.cell(row_sr, 8).value = sr
            ws_d.cell(row_d, 8).value = d
            ws_o.cell(row_o, 8).value = open
        elif 'RenderServer' == module:
            ws_sr.cell(row_sr, 9).value = sr
            ws_d.cell(row_d, 9).value = d
            ws_o.cell(row_o, 9).value = open
        elif 'Review' == module:
            ws_sr.cell(row_sr, 10).value = sr
            ws_d.cell(row_d, 10).value = d
            ws_o.cell(row_o, 10).value = open

    wb_bd.save(dst)


def match_today(ws):
    today = datetime.datetime.today().date()
    for i in range(1, ws.max_row +1):
        if ws.cell(i, 1).value is None:
            continue

        date = openpyxl.utils.datetime.from_excel(ws.cell(i, 1).value).date()
        if today == date:
            return i


if __name__ == '__main__':
    if len(sys.argv) == 3:
        update_burn_down_xlsx(sys.argv[1], sys.argv[2])
    else:
        update_burn_down_xlsx('E:/Code/CSharp/TFSCapture/MailSync/temp(1).xlsx', 'E:/Documents/项目/大C/PM/BugBurndown.xlsx')
